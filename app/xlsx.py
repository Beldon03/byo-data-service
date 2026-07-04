import io
from datetime import date, datetime, time
from typing import Any

import openpyxl

from app.ingestion import CsvError


def parse_xlsx(data: bytes) -> tuple[list[str], list[list[str]]]:
    """Extract the active sheet as header + string rows for the shared
    ingestion pipeline (sanitization, type inference, DDL, insert)."""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        # openpyxl surfaces corrupt input through many exception types
        # (BadZipFile, InvalidFileException, KeyError, ...).
        raise CsvError("file is not a valid XLSX workbook") from exc
    try:
        # The active tab can be a chartsheet, which holds no cells; prefer it
        # when usable, otherwise fall back to the first real worksheet.
        # (openpyxl 3.1.x often fails to load chartsheet-bearing workbooks at
        # all; that surfaces through the except above as a 400.)
        candidates = (workbook.active, *workbook.worksheets)
        sheet = next((s for s in candidates if s is not None and hasattr(s, "iter_rows")), None)
        if sheet is None:
            raise CsvError("XLSX workbook has no data sheets")
        raw = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    # Excel files routinely carry phantom trailing rows and columns from
    # stray formatting; trim to the rectangle that actually holds values.
    width = max((_used_width(row) for row in raw), default=0)
    rows = [
        [_cell_to_text(value) for value in row[:width]] + [""] * (width - len(row))
        for row in raw
    ]
    rows = [row for row in rows if any(cell != "" for cell in row)]

    if not rows:
        raise CsvError("XLSX sheet is empty")
    if len(rows) == 1:
        raise CsvError("XLSX sheet contains a header row but no data rows")
    return rows[0], rows[1:]


def _used_width(row: list[Any]) -> int:
    width = 0
    for index, value in enumerate(row):
        if value is not None and str(value).strip() != "":
            width = index + 1
    return width


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        # Excel represents pure dates as midnight datetimes.
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date | time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # Excel stores every number as a float; keep 3 from becoming "3.0"
        # so integer columns still infer as integer.
        return str(int(value))
    return str(value)
